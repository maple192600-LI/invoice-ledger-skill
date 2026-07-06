"""Excel draft ledger writer for the invoice draft ledger pipeline."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from enum import Enum
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
import yaml

from ..contracts import CorrectionStatus, LedgerRow, WriteAction, WriteResult, normalize_date


def _load_mapping(mapping_path: str | Path) -> dict[str, Any]:
    with Path(mapping_path).open("r", encoding="utf-8") as file:
        loaded = yaml.safe_load(file)
    if not isinstance(loaded, dict):
        raise ValueError("Invalid draft ledger mapping.")
    return loaded


def _load_status_labels(mapping_path: str | Path) -> dict[str, str]:
    status_path = Path(mapping_path).parent / "status_messages.yaml"
    if not status_path.exists():
        return {}
    with status_path.open("r", encoding="utf-8") as file:
        loaded = yaml.safe_load(file)
    statuses = loaded.get("statuses", {}) if isinstance(loaded, dict) else {}
    return {
        str(status): str(spec.get("zh"))
        for status, spec in statuses.items()
        if isinstance(spec, dict) and spec.get("zh")
    }


def _headers_by_name(ws) -> dict[str, int]:
    return {
        str(ws.cell(1, column).value).strip(): column
        for column in range(1, ws.max_column + 1)
        if ws.cell(1, column).value is not None
    }


def _field_columns_from_headers(mapping: dict[str, Any], headers: dict[str, int]) -> dict[str, int]:
    columns: dict[str, int] = {}
    for field_name, spec in mapping["fields"].items():
        if not isinstance(spec, dict):
            continue
        for header in spec.get("headers", []):
            if header in headers:
                columns[str(field_name)] = headers[header]
                break
    return columns


CELL_REFERENCE_RE = re.compile(r"(?<![A-Za-z0-9_])(\$?)([A-Z]{1,3})(\$?)(\d+)(?![A-Za-z0-9_])")


def _shift_formula_columns_for_insert(formula: str, insert_column: int, amount: int = 1) -> str:
    def replace_cell_reference(match: re.Match[str]) -> str:
        column_absolute, column_name, row_absolute, row_number = match.groups()
        column_index = column_index_from_string(column_name)
        if column_index < insert_column:
            return match.group(0)
        shifted_column = get_column_letter(column_index + amount)
        return f"{column_absolute}{shifted_column}{row_absolute}{row_number}"

    return CELL_REFERENCE_RE.sub(replace_cell_reference, formula)


def _shift_formulas_for_inserted_columns(ws, insert_column: int, amount: int = 1) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = _shift_formula_columns_for_insert(cell.value, insert_column, amount)


def _create_header(ws, header: str, target_column: int) -> None:
    if target_column <= ws.max_column:
        ws.insert_cols(target_column)
        _shift_formulas_for_inserted_columns(ws, target_column)
    ws.cell(1, target_column).value = header


def _ensure_configured_headers(ws, mapping: dict[str, Any]) -> None:
    headers = _headers_by_name(ws)
    for spec in mapping["fields"].values():
        if not isinstance(spec, dict) or spec.get("create_if_missing") is not True:
            continue
        header_names = spec.get("headers", [])
        if not header_names or any(header in headers for header in header_names):
            continue
        header = str(header_names[0])
        target_column = ws.max_column + 1
        insert_after = spec.get("insert_after")
        if isinstance(insert_after, str):
            anchor_column = _field_columns_from_headers(mapping, headers).get(insert_after)
            if anchor_column is not None:
                target_column = anchor_column + 1
        _create_header(ws, header, target_column)
        headers = _headers_by_name(ws)


def _field_columns(ws, mapping: dict[str, Any]) -> dict[str, int]:
    _ensure_configured_headers(ws, mapping)
    headers = _headers_by_name(ws)
    columns: dict[str, int] = {}
    for field_name, spec in mapping["fields"].items():
        for header in spec.get("headers", []):
            if header in headers:
                columns[field_name] = headers[header]
                break
    return columns


def _field_specs(mapping: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {
        str(field_name): spec
        for field_name, spec in mapping["fields"].items()
        if isinstance(spec, dict)
    }


def _validate_required_headers(mapping: dict[str, Any], columns: dict[str, int]) -> None:
    missing = [
        field_name
        for field_name, spec in mapping["fields"].items()
        if spec.get("required") is True and field_name not in columns
    ]
    if missing:
        raise ValueError("Missing required header mapping for required header fields: " + ", ".join(missing))


def _date_value(value: Any) -> date | Any:
    if value in {None, ""}:
        return value
    try:
        normalized = normalize_date(value)
    except ValueError:
        return value
    return date.fromisoformat(normalized)


def _decimal_value(value: Decimal) -> int | float:
    if value == value.to_integral_value():
        return int(value)
    return float(value)


def _cell_value(field_name: str, value: Any, status_labels: dict[str, str]) -> Any:
    if field_name in {"processed_at", "invoice_date", "correction_time"}:
        return _date_value(value)
    if field_name == "recognition_status":
        status = value.value if isinstance(value, Enum) else str(value)
        return status_labels.get(status, status)
    if isinstance(value, Enum):
        return value.value
    if isinstance(value, Decimal):
        return _decimal_value(value)
    if isinstance(value, (int, float, bool)):
        return value
    if isinstance(value, list):
        return ",".join(str(item) for item in value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    return str(value)


def _formula_value(template: str, row_number: int, columns: dict[str, int]) -> str | None:
    tokens = re.findall(r"{([^}]+)}", template)
    if any(token not in columns for token in tokens):
        return None
    formula = template
    for token in tokens:
        formula = formula.replace(f"{{{token}}}", f"{get_column_letter(columns[token])}{row_number}")
    return "=" + formula.lstrip("=")


def _mapped_cell_value(
    field_name: str,
    spec: dict[str, Any],
    values: dict[str, Any],
    row_number: int,
    columns: dict[str, int],
    status_labels: dict[str, str],
) -> Any:
    if spec.get("value") == "row_number":
        return row_number - 1
    formula_template = spec.get("formula")
    if isinstance(formula_template, str):
        formula_row_type = spec.get("formula_when_row_type")
        if formula_row_type is None or values.get("row_type") == formula_row_type:
            formula = _formula_value(formula_template, row_number, columns)
            if formula is not None:
                return formula
    return _cell_value(field_name, values.get(field_name), status_labels)


MONEY_FIELD_NAMES = {
    "line_amount",
    "line_tax_amount",
    "line_total_with_tax",
    "invoice_amount_total",
    "invoice_tax_total",
    "invoice_total_with_tax",
}


def _apply_cell_format(cell, field_name: str, value: Any) -> None:
    if field_name in {"processed_at", "invoice_date", "correction_time"} and value not in {None, ""}:
        cell.number_format = "yyyy-mm-dd"
    if field_name in MONEY_FIELD_NAMES and value not in {None, ""}:
        cell.number_format = "0.00"


def _row_dict(row: LedgerRow) -> dict[str, Any]:
    return row.model_dump(mode="python")


def _normalize_existing(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, (int, float, Decimal)):
        decimal = Decimal(str(value))
        return f"{decimal.quantize(Decimal('0.01')):.2f}"
    return str(value).strip()


def _dedupe_key_from_values(values: dict[str, Any]) -> str:
    key_fields = [
        "invoice_line_key",
        "source_file",
        "invoice_no",
        "invoice_date",
        "item_name",
        "line_amount",
        "line_tax_amount",
        "invoice_total_with_tax",
    ]
    present = [str(values.get("invoice_line_key") or "")]
    if present[0]:
        return present[0]
    return "|".join(_normalize_existing(values.get(field)) for field in key_fields[1:])


def _existing_values(ws, row_number: int, columns: dict[str, int]) -> dict[str, Any]:
    return {
        field: ws.cell(row=row_number, column=column).value
        for field, column in columns.items()
    }


def _existing_dedupe_keys(ws, columns: dict[str, int]) -> set[str]:
    keys: set[str] = set()
    for row_number in range(2, ws.max_row + 1):
        values = _existing_values(ws, row_number, columns)
        key = _dedupe_key_from_values(values)
        if key.strip("|"):
            keys.add(key)
    return keys


def _dedupe_values_for_available_columns(values: dict[str, Any], columns: dict[str, int]) -> dict[str, Any]:
    return {field: values.get(field) for field in columns}


def _find_existing_row(ws, columns: dict[str, int], key: str) -> int | None:
    for row_number in range(2, ws.max_row + 1):
        values = _existing_values(ws, row_number, columns)
        if _dedupe_key_from_values(values) == key:
            return row_number
    return None


def _find_existing_row_by_field(ws, columns: dict[str, int], field_name: str, value: Any) -> int | None:
    column = columns.get(field_name)
    if column is None or value in {None, ""}:
        return None
    normalized_target = _normalize_existing(value)
    for row_number in range(2, ws.max_row + 1):
        if _normalize_existing(ws.cell(row=row_number, column=column).value) == normalized_target:
            return row_number
    return None


def _last_data_row(ws) -> int:
    for row_number in range(ws.max_row, 1, -1):
        if any(
            ws.cell(row=row_number, column=column).value not in {None, ""}
            for column in range(1, ws.max_column + 1)
        ):
            return row_number
    return 1


def write_ledger_rows(
    workbook_path: str | Path,
    target_sheet: str,
    ledger_rows: list[LedgerRow],
    mapping_path: str | Path,
    run_id: str,
    update_existing: bool = False,
) -> WriteResult:
    mapping = _load_mapping(mapping_path)
    field_specs = _field_specs(mapping)
    status_labels = _load_status_labels(mapping_path)
    workbook = load_workbook(workbook_path)
    if target_sheet not in workbook.sheetnames:
        raise ValueError(f"Target sheet not found: {target_sheet}")

    ws = workbook[target_sheet]
    columns = _field_columns(ws, mapping)
    _validate_required_headers(mapping, columns)
    existing_keys = _existing_dedupe_keys(ws, columns)
    result = WriteResult(run_id=run_id, target_sheet=target_sheet)

    for row in ledger_rows:
        values = _row_dict(row)
        comparable_values = _dedupe_values_for_available_columns(values, columns)
        key = _dedupe_key_from_values(comparable_values)
        existing_row = None
        if update_existing:
            existing_row = _find_existing_row_by_field(ws, columns, "draft_row_id", values.get("draft_row_id"))
        if existing_row is None:
            existing_row = _find_existing_row(ws, columns, key)
        if existing_row is not None and update_existing:
            values["correction_status"] = CorrectionStatus.CORRECTED
            for field, column in columns.items():
                cell_value = _mapped_cell_value(
                    field,
                    field_specs.get(field, {}),
                    values,
                    existing_row,
                    columns,
                    status_labels,
                )
                cell = ws.cell(row=existing_row, column=column)
                cell.value = cell_value
                _apply_cell_format(cell, field, cell_value)
            result.updated_rows += 1
            result.actions.append(
                {
                    "action": WriteAction.UPDATED.value,
                    "excel_row": existing_row,
                    "invoice_line_key": row.invoice_line_key,
                }
            )
            continue

        if existing_row is not None and not update_existing:
            result.skipped_duplicate_rows += 1
            result.actions.append(
                {
                    "action": WriteAction.SKIPPED_DUPLICATE.value,
                    "invoice_line_key": row.invoice_line_key,
                    "source_file": row.source_file,
                    "invoice_no": row.invoice_no,
                    "message": (
                        f"疑似重复：文件 {Path(row.source_file).name}，发票号码：{row.invoice_no or '未识别'}，本次未写入。"
                    ),
                }
            )
            continue

        target_row = _last_data_row(ws) + 1
        for field, column in columns.items():
            cell_value = _mapped_cell_value(
                field,
                field_specs.get(field, {}),
                values,
                target_row,
                columns,
                status_labels,
            )
            cell = ws.cell(row=target_row, column=column)
            cell.value = cell_value
            _apply_cell_format(cell, field, cell_value)
        existing_keys.add(key)
        result.added_rows += 1
        result.actions.append(
            {
                "action": WriteAction.ADDED.value,
                "excel_row": target_row,
                "invoice_line_key": row.invoice_line_key,
            }
        )

    workbook.save(workbook_path)
    return result
