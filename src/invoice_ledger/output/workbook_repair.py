"""加载 xlsx 前的样式容错修复。

WPS Office 等工具保存 xlsx 时,可能在 ``xl/styles.xml`` 写入空或畸形
的 ``<fill>`` 元素,导致 openpyxl ``load_workbook`` 抛
``TypeError: expected Fill``(openpyxl 官方 issue #1805)。

本模块在加载前把这类没有 ``patternFill`` / ``gradientFill`` 子元素的
``<fill>`` 补成合法的 ``<patternFill patternType="none"/>``,使加载不再
因样式崩溃。修复只发生在临时副本上,原文件保持不动。
"""

from __future__ import annotations

import os
import tempfile
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook as _openpyxl_load_workbook

_SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_FILL = f"{{{_SPREADSHEET_NS}}}fill"
_PATTERN_FILL = f"{{{_SPREADSHEET_NS}}}patternFill"
_GRADIENT_FILL = f"{{{_SPREADSHEET_NS}}}gradientFill"
_FILLS = f"{{{_SPREADSHEET_NS}}}fills"
_STYLES_ENTRY = "xl/styles.xml"


def repair_styles_xml(xml_bytes: bytes) -> bytes | None:
    """补全 styles.xml 里没有 patternFill/gradientFill 子元素的空 <fill>。

    返回修复后的字节;若没有任何坏 fill 需要修复,返回 None。
    """
    text = xml_bytes.decode("utf-8", errors="replace")
    declaration = ""
    body = text
    leading = len(text) - len(text.lstrip())
    if text[leading:].startswith("<?xml"):
        end = text.index("?>", leading) + 2
        declaration = text[:end]
        body = text[end:]

    root = ET.fromstring(body)
    fills_element = root.find(_FILLS)
    if fills_element is None:
        return None

    repaired = 0
    for fill in list(fills_element):
        if fill.tag != _FILL:
            continue
        if any(child.tag in (_PATTERN_FILL, _GRADIENT_FILL) for child in fill):
            continue
        for child in list(fill):
            fill.remove(child)
        ET.SubElement(fill, _PATTERN_FILL).set("patternType", "none")
        repaired += 1

    if repaired == 0:
        return None

    ET.register_namespace("", _SPREADSHEET_NS)
    return (declaration + ET.tostring(root, encoding="unicode")).encode("utf-8")


def make_loadable_copy(source_path: str | Path) -> Path:
    """返回 openpyxl 可安全加载的 xlsx 路径。

    样式干净时原样返回源路径;含坏 fill 时写一份修复后的临时副本。
    """
    source = Path(source_path)
    with zipfile.ZipFile(source, "r") as archive:
        names = archive.namelist()
        styles = archive.read(_STYLES_ENTRY) if _STYLES_ENTRY in names else None
    if styles is None:
        return source
    repaired = repair_styles_xml(styles)
    if repaired is None:
        return source

    descriptor, temporary_name = tempfile.mkstemp(suffix=".repaired.xlsx")
    os.close(descriptor)
    temporary = Path(temporary_name)
    with zipfile.ZipFile(source, "r") as archive, \
         zipfile.ZipFile(temporary, "w", zipfile.ZIP_DEFLATED) as rebuilt:
        for item in archive.infolist():
            data = archive.read(item.filename)
            if item.filename == _STYLES_ENTRY:
                data = repaired
            rebuilt.writestr(item, data)
    return temporary


def safe_load_workbook(path: str | Path, **kwargs):
    """加载台账;遇到 WPS 写的坏 fill 自动修复后加载。

    原文件不被修改(修复只发生在临时副本上,加载后即删除)。
    """
    loadable = make_loadable_copy(path)
    if loadable == Path(path):
        return _openpyxl_load_workbook(loadable, **kwargs)
    try:
        return _openpyxl_load_workbook(loadable, **kwargs)
    finally:
        loadable.unlink(missing_ok=True)
