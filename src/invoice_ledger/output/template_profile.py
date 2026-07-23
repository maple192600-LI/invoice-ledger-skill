"""Template profile loading and workbook compatibility checks."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import yaml


def load_template_profile(path: str | Path) -> dict[str, Any]:
    profile_path = Path(path)
    if not profile_path.exists():
        raise FileNotFoundError(f"Template profile not found: {profile_path}")
    with profile_path.open("r", encoding="utf-8") as file:
        loaded = yaml.safe_load(file)
    if not isinstance(loaded, dict):
        raise ValueError(f"Invalid template profile: {profile_path}")
    if not isinstance(loaded.get("sheets"), dict) or not loaded["sheets"]:
        raise ValueError(f"Template profile has no sheets: {profile_path}")
    return loaded


def _headers(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for column in range(1, ws.max_column + 1):
        value = ws.cell(1, column).value
        if value is None:
            continue
        header = str(value).strip()
        if header not in headers:
            headers[header] = column
    return headers


def _duplicate_headers(ws) -> list[dict[str, Any]]:
    columns_by_header: dict[str, list[int]] = {}
    for column in range(1, ws.max_column + 1):
        value = ws.cell(1, column).value
        if value is None:
            continue
        header = str(value).strip()
        columns_by_header.setdefault(header, []).append(column)
    return [
        {"header": header, "columns": columns}
        for header, columns in sorted(columns_by_header.items())
        if len(columns) > 1
    ]


def _field_columns(fields: dict[str, Any], headers: dict[str, int]) -> dict[str, str]:
    columns: dict[str, str] = {}
    for field_name, spec in fields.items():
        if not isinstance(spec, dict):
            continue
        for header in spec.get("headers", []):
            if header in headers:
                columns[str(field_name)] = str(header)
                break
    return columns


def _missing_required(fields: dict[str, Any], field_columns: dict[str, str]) -> list[dict[str, Any]]:
    missing: list[dict[str, Any]] = []
    for field_name, spec in fields.items():
        if not isinstance(spec, dict) or spec.get("required") is not True:
            continue
        if field_name in field_columns:
            continue
        missing.append(
            {
                "field": str(field_name),
                "headers": [str(header) for header in spec.get("headers", [])],
            }
        )
    return missing


def validate_template_workbook(workbook_path: str | Path, profile: dict[str, Any]) -> dict[str, Any]:
    path = Path(workbook_path)
    workbook = load_workbook(path, data_only=False)
    try:
        report: dict[str, Any] = {
            "status": "passed",
            "template_id": profile.get("template_id"),
            "template_version": profile.get("template_version"),
            "workbook": str(path),
            "sheets": {},
            "failure_reasons": [],
            "blocked_write": False,
            "recommended_action": "none",
        }
        for sheet_key, sheet_spec in profile["sheets"].items():
            sheet_name = str(sheet_spec.get("name", ""))
            fields = sheet_spec.get("fields", {})
            if sheet_name not in workbook.sheetnames:
                report["sheets"][sheet_key] = {
                    "sheet_name": sheet_name,
                    "status": "failed",
                    "missing_sheet": True,
                    "missing_required_fields": [],
                    "duplicate_headers": [],
                    "field_columns": {},
                    "extra_headers": [],
                }
                report["status"] = "failed"
                if "missing_sheet" not in report["failure_reasons"]:
                    report["failure_reasons"].append("missing_sheet")
                continue

            ws = workbook[sheet_name]
            headers = _headers(ws)
            duplicates = _duplicate_headers(ws)
            field_columns = _field_columns(fields, headers)
            configured_headers = {
                str(header)
                for spec in fields.values()
                if isinstance(spec, dict)
                for header in spec.get("headers", [])
            }
            missing = _missing_required(fields, field_columns)
            sheet_status = "failed" if missing or duplicates else "passed"
            if missing:
                report["status"] = "failed"
                if "missing_required_field" not in report["failure_reasons"]:
                    report["failure_reasons"].append("missing_required_field")
            if duplicates:
                report["status"] = "failed"
                if "duplicate_header" not in report["failure_reasons"]:
                    report["failure_reasons"].append("duplicate_header")
            report["sheets"][sheet_key] = {
                "sheet_name": sheet_name,
                "status": sheet_status,
                "missing_sheet": False,
                "missing_required_fields": missing,
                "duplicate_headers": duplicates,
                "field_columns": field_columns,
                "extra_headers": sorted(set(headers) - configured_headers),
            }
        if report["status"] != "passed":
            report["blocked_write"] = True
            report["recommended_action"] = "update_template_or_profile_before_write"
        return report
    finally:
        workbook.close()
