"""Profile-driven Excel writer for invoice collection templates."""

from __future__ import annotations

from contextlib import contextmanager
from copy import copy
from datetime import date, datetime
from decimal import Decimal
from enum import Enum
import errno
from hashlib import sha1
import os
from pathlib import Path
import tempfile
from typing import Any
from unicodedata import east_asian_width

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import yaml

from ..contracts import LedgerRow, RecognitionNotice, RecognitionStatus, WriteAction, WriteResult, normalize_date
from ..parsing.invoice_identity import has_standard_digital_invoice_number
from .._paths import PROJECT_ROOT
from ..validation.review_notes import user_review_remark
from . import duplicate_rows
from .recognition_notices import duplicate_notice_from_ledger_row
from .template_profile import load_template_profile, validate_template_workbook


REVIEW_REMARK_PREFIXES = (
    "missing evidence ",
    "missing ",
    "conflict ",
    "low confidence ",
    "amount_total + tax_total",
    "incomplete amount breakdown",
    "sum line_amount",
    "sum line_tax_amount",
    "line ",
    "digital invoice ",
    "failed",
    "error",
    "待复核：",
    "需复核：",
)

DEFAULT_STATUS_LABELS = {
    RecognitionStatus.READY.value: RecognitionStatus.READY.value,
    RecognitionStatus.REVIEW_REQUIRED.value: RecognitionStatus.REVIEW_REQUIRED.value,
    RecognitionStatus.UNMODELED.value: RecognitionStatus.UNMODELED.value,
    RecognitionStatus.FAILED.value: RecognitionStatus.FAILED.value,
}

ROW_FINGERPRINT_FIELDS = (
    "invoice_code",
    "invoice_no",
    "digital_invoice_no",
    "invoice_date",
    "seller_tax_id",
    "buyer_tax_id",
    "item_name",
    "line_amount",
    "line_tax_amount",
    "line_total_with_tax",
    "invoice_total_with_tax",
)

MIN_COLUMN_WIDTH = 6.0
MAX_COLUMN_WIDTH = 255.0


@contextmanager
def _exclusive_workbook_lock(workbook_path: Path):
    lock_root = Path(tempfile.gettempdir()) / "invoice_ledger_locks"
    lock_root.mkdir(exist_ok=True)
    lock_name = sha1(str(workbook_path.resolve()).casefold().encode("utf-8")).hexdigest()
    with (lock_root / lock_name).open("a+b") as lock_file:
        if lock_file.seek(0, os.SEEK_END) == 0:
            lock_file.write(b"\0")
            lock_file.flush()
        lock_file.seek(0)
        if os.name == "nt":
            import msvcrt

            while True:
                try:
                    msvcrt.locking(lock_file.fileno(), msvcrt.LK_LOCK, 1)
                    break
                except OSError as exc:
                    if exc.errno not in {errno.EACCES, errno.EDEADLK}:
                        raise
            try:
                yield
            finally:
                lock_file.seek(0)
                msvcrt.locking(lock_file.fileno(), msvcrt.LK_UNLCK, 1)
        else:
            import fcntl

            fcntl.flock(lock_file.fileno(), fcntl.LOCK_EX)
            try:
                yield
            finally:
                fcntl.flock(lock_file.fileno(), fcntl.LOCK_UN)


def _status_labels() -> dict[str, str]:
    status_path = PROJECT_ROOT / "config" / "status_messages.yaml"
    if not status_path.exists():
        return DEFAULT_STATUS_LABELS
    with status_path.open("r", encoding="utf-8") as file:
        loaded = yaml.safe_load(file)
    statuses = loaded.get("statuses", {}) if isinstance(loaded, dict) else {}
    labels = {
        str(status): str(spec.get("zh"))
        for status, spec in statuses.items()
        if isinstance(spec, dict) and spec.get("zh")
    }
    return DEFAULT_STATUS_LABELS | labels


def _headers(ws) -> dict[str, int]:
    return {
        str(ws.cell(1, column).value).strip(): column
        for column in range(1, ws.max_column + 1)
        if ws.cell(1, column).value is not None
    }


def _field_columns(fields: dict[str, Any], headers: dict[str, int]) -> dict[str, int]:
    columns: dict[str, int] = {}
    for field_name, spec in fields.items():
        if not isinstance(spec, dict):
            continue
        for header in spec.get("headers", []):
            if header in headers:
                columns[str(field_name)] = headers[header]
                break
    return columns


def _date_value(value: Any) -> date | Any:
    if value in {None, ""}:
        return value
    try:
        normalized = normalize_date(value)
    except ValueError:
        return value
    return date.fromisoformat(normalized) if normalized else value


_DATETIME_FORMATS = ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M")


def _datetime_value(value: Any) -> datetime | Any:
    if value in {None, ""}:
        return value
    text = str(value).strip()
    for fmt in _DATETIME_FORMATS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return _date_value(value)


def _cell_value(field_name: str, value: Any) -> Any:
    if field_name == "processed_at":
        return _datetime_value(value)
    if field_name in {"invoice_date", "correction_time"}:
        return _date_value(value)
    if isinstance(value, Enum):
        value = value.value
    if field_name in {"recognition_status", "issue_type"}:
        return _status_labels().get(str(value), str(value))
    if isinstance(value, Decimal):
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value
    if isinstance(value, list):
        return ",".join(str(item) for item in value)
    return value


def _set_cell_value(cell, value: Any) -> None:
    cell.value = value
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        cell.data_type = "s"


def _remark_parts(remark: str | None) -> list[str]:
    text = str(remark or "").strip()
    if not text:
        return []
    return [part.strip() for part in text.replace(";", "；").split("；") if part.strip()]


def _is_review_remark_part(part: str) -> bool:
    return part.startswith(REVIEW_REMARK_PREFIXES) or "不一致" in part or "缺失" in part


def _split_remark(remark: str | None) -> tuple[list[str], list[str]]:
    review_parts: list[str] = []
    business_parts: list[str] = []
    for part in _remark_parts(remark):
        if _is_review_remark_part(part):
            review_parts.append(part)
        else:
            business_parts.append(part)
    return review_parts, business_parts


def _display_remark(row: LedgerRow) -> str:
    if row.review_remark or row.context_remark:
        display_parts = []
        review_text = user_review_remark(row.review_remark) if row.review_remark else ""
        if review_text:
            display_parts.append(review_text)
        display_parts.extend(_remark_parts(row.context_remark))
        return "；".join(display_parts)

    if row.recognition_status == RecognitionStatus.READY:
        return "；".join(_remark_parts(row.remark))

    review_parts, business_parts = _split_remark(row.remark)
    display_parts = []
    review_text = user_review_remark("；".join(review_parts)) if review_parts else ""
    if review_text:
        display_parts.append(review_text)
    display_parts.extend(business_parts)
    return "；".join(display_parts)


def _is_digital_vat_invoice(row: LedgerRow) -> bool:
    return has_standard_digital_invoice_number(
        row.schema_id,
        row.variant_id,
        row.invoice_no,
    )


def _row_values(row: LedgerRow) -> dict[str, Any]:
    values = row.model_dump(mode="python")
    display_remark = _display_remark(row)
    if _is_digital_vat_invoice(row):
        values["digital_invoice_no"] = values.get("invoice_no")
        values["invoice_no"] = None
    else:
        values["digital_invoice_no"] = None
    values["issue_type"] = row.recognition_status
    values["remark"] = display_remark
    return values


def _value_for_field(
    field_name: str,
    spec: dict[str, Any],
    values: dict[str, Any],
    target_row: int,
) -> Any:
    if spec.get("value") == "row_number":
        return target_row - 1
    if "static" in spec:
        return spec["static"]
    source = str(spec.get("source") or field_name)
    return _cell_value(field_name, values.get(source))


def _append_rows(ws, fields: dict[str, Any], rows: list[LedgerRow]) -> tuple[int, int, list[tuple[LedgerRow, dict[str, Any]]]]:
    columns = _field_columns(fields, _headers(ws))
    written = 0
    skipped = 0
    skipped_rows: list[tuple[LedgerRow, dict[str, Any]]] = []
    existing_draft_row_ids = duplicate_rows.existing_draft_row_ids(ws, columns)
    existing_invoice_line_keys = duplicate_rows.existing_invoice_line_keys(ws, fields, columns)
    existing_row_fingerprints = duplicate_rows.existing_row_fingerprints(ws, columns, ROW_FINGERPRINT_FIELDS)
    existing_draft_row_id_rows = duplicate_rows.existing_draft_row_id_rows(ws, columns)
    existing_invoice_line_key_rows = duplicate_rows.existing_invoice_line_key_rows(ws, fields, columns)
    existing_row_fingerprint_rows = duplicate_rows.existing_row_fingerprint_rows(
        ws,
        columns,
        ROW_FINGERPRINT_FIELDS,
    )
    weak_duplicate_rows: list[tuple[LedgerRow, dict[str, Any]]] = []
    for row in rows:
        values = _row_values(row)
        row_fingerprint = duplicate_rows.row_fingerprint_from_values(
            fields,
            columns,
            values,
            ROW_FINGERPRINT_FIELDS,
            _value_for_field,
        )
        weak_identity = not row.invoice_no and not _is_digital_vat_invoice(row)
        is_duplicate = (
            row.draft_row_id in existing_draft_row_ids
            or row.invoice_line_key in existing_invoice_line_keys
            or (row_fingerprint is not None and row_fingerprint in existing_row_fingerprints)
        )
        if is_duplicate:
            existing_row_number = (
                existing_draft_row_id_rows.get(row.draft_row_id)
                or existing_invoice_line_key_rows.get(row.invoice_line_key)
                or (existing_row_fingerprint_rows.get(row_fingerprint) if row_fingerprint is not None else None)
                or duplicate_rows.first_row_by_invoice_number(ws, columns, row.invoice_no)
            )
            existing_context = duplicate_rows.existing_row_context(ws, columns, existing_row_number)
            if weak_identity:
                row.context_remark = "；".join(
                    part for part in [row.context_remark, "疑似重复（弱身份票），请人工确认"] if part
                )
                weak_duplicate_rows.append((row, existing_context))
            else:
                skipped += 1
                skipped_rows.append((row, existing_context))
                continue
        target_row = _last_data_row(ws) + 1
        for field_name, column in columns.items():
            spec = fields.get(field_name, {})
            cell = ws.cell(target_row, column)
            _set_cell_value(
                cell,
                _value_for_field(
                    field_name,
                    spec if isinstance(spec, dict) else {},
                    values,
                    target_row,
                ),
            )
            if isinstance(cell.value, datetime):
                cell.number_format = "YYYY-MM-DD HH:MM"
        existing_draft_row_ids.add(row.draft_row_id)
        existing_invoice_line_keys.add(row.invoice_line_key)
        written += 1
    return written, skipped, skipped_rows, weak_duplicate_rows


def _notice_values(notice: RecognitionNotice) -> dict[str, Any]:
    values = notice.model_dump(mode="python")
    values["draft_row_id"] = notice.notice_id
    values["invoice_line_key"] = notice.notice_id
    return values


def _value_for_notice_field(
    field_name: str,
    spec: dict[str, Any],
    values: dict[str, Any],
    target_row: int,
) -> Any:
    if spec.get("value") == "row_number":
        return target_row - 1
    source = str(spec.get("source") or field_name)
    if source in values:
        return _cell_value(field_name, values.get(source))
    if field_name in values:
        return _cell_value(field_name, values.get(field_name))
    if "static" in spec:
        return spec["static"]
    return None


def _append_notice_rows(
    ws,
    fields: dict[str, Any],
    notices: list[RecognitionNotice],
) -> tuple[int, int]:
    columns = _field_columns(fields, _headers(ws))
    written = 0
    skipped = 0
    existing_notice_ids = duplicate_rows.existing_draft_row_ids(ws, columns)
    for notice in notices:
        values = _notice_values(notice)
        if notice.notice_id in existing_notice_ids:
            skipped += 1
            continue
        target_row = _last_data_row(ws) + 1
        for field_name, column in columns.items():
            spec = fields.get(field_name, {})
            cell = ws.cell(target_row, column)
            _set_cell_value(
                cell,
                _value_for_notice_field(
                    field_name,
                    spec if isinstance(spec, dict) else {},
                    values,
                    target_row,
                ),
            )
            if isinstance(cell.value, datetime):
                cell.number_format = "YYYY-MM-DD HH:MM"
        existing_notice_ids.add(notice.notice_id)
        written += 1
    return written, skipped


def _last_data_row(ws) -> int:
    for row_number in range(ws.max_row, 1, -1):
        if any(
            ws.cell(row_number, column).value not in {None, ""}
            for column in range(1, ws.max_column + 1)
        ):
            return row_number
    return 1


def _display_width(value: Any) -> int:
    if value in {None, ""}:
        return 0
    if isinstance(value, datetime):
        text = value.strftime("%Y-%m-%d %H:%M")
    elif isinstance(value, date):
        text = value.isoformat()
    else:
        text = str(value)
    return max(
        (
            sum(2 if east_asian_width(char) in {"W", "F"} else 1 for char in line)
            for line in text.splitlines()
        ),
        default=0,
    )


def _repair_plain_headers(ws) -> None:
    styled_header = next(
        (
            cell
            for cell in ws[1]
            if cell.value not in {None, ""}
            and cell.fill.fill_type
            and cell.font.bold
        ),
        None,
    )
    if styled_header is None:
        return
    for cell in ws[1]:
        if (
            cell.value not in {None, ""}
            and not cell.fill.fill_type
            and not cell.font.bold
        ):
            cell._style = copy(styled_header._style)


def _autofit_collection_sheet(ws) -> None:
    last_row = _last_data_row(ws)
    for column in range(1, ws.max_column + 1):
        content_width = max(
            _display_width(ws.cell(row, column).value)
            for row in range(1, last_row + 1)
        )
        width = min(MAX_COLUMN_WIDTH, max(MIN_COLUMN_WIDTH, content_width + 2))
        ws.column_dimensions[get_column_letter(column)].width = width
        for row in range(2, last_row + 1):
            cell = ws.cell(row, column)
            if cell.alignment.wrap_text is True:
                alignment = copy(cell.alignment)
                alignment.wrap_text = False
                cell.alignment = alignment
                ws.row_dimensions[row].height = None


def _summary_rows(rows: list[LedgerRow]) -> list[LedgerRow]:
    summaries: dict[str, LedgerRow] = {}
    for row in rows:
        if row.invoice_key not in summaries:
            summaries[row.invoice_key] = row
        if row.row_type == "汇总":
            summaries[row.invoice_key] = row
    return list(summaries.values())


def _issue_rows(rows: list[LedgerRow]) -> list[LedgerRow]:
    return [
        row
        for row in rows
        if row.recognition_status != RecognitionStatus.READY
        or bool(row.review_remark)
    ]


def _write_with_template_profile(
    workbook_path: str | Path,
    template_profile_path: str | Path,
    ledger_rows: list[LedgerRow],
    run_id: str,
    recognition_notices: list[RecognitionNotice] | None = None,
) -> WriteResult:
    profile = load_template_profile(template_profile_path)
    drift_report = validate_template_workbook(workbook_path, profile)
    if drift_report.get("blocked_write") is True or drift_report["status"] != "passed":
        raise ValueError(
            {
                "message": "Template workbook does not match profile",
                "template_drift_report": drift_report,
            }
        )

    workbook_path = Path(workbook_path)
    workbook = load_workbook(workbook_path)
    result = WriteResult(run_id=run_id, target_sheet=str(profile.get("template_id") or "template"))
    temporary_path: Path | None = None
    try:
        for sheet_key, sheet_spec in profile["sheets"].items():
            ws = workbook[str(sheet_spec["name"])]
            _repair_plain_headers(ws)
            fields = sheet_spec.get("fields", {})
            mode = sheet_spec.get("mode")
            written = 0
            skipped = 0
            if mode == "ledger_rows":
                written, skipped, skipped_rows, weak_duplicates = _append_rows(ws, fields, ledger_rows)
                if recognition_notices is not None:
                    existing_notice_ids = {notice.notice_id for notice in recognition_notices}
                    for skipped_row, existing_row in skipped_rows:
                        notice = duplicate_notice_from_ledger_row(skipped_row, existing_row)
                        if notice.notice_id in existing_notice_ids:
                            continue
                        recognition_notices.append(notice)
                        existing_notice_ids.add(notice.notice_id)
                        duplicate_position = (
                            f"重复位置：采集表第 {existing_row['excel_row']} 行；"
                            if existing_row and existing_row.get("excel_row")
                            else ""
                        )
                        invoice_no = notice.invoice_no or "发票号码未识别"
                        result.messages.append(
                            f"疑似重复：文件 {notice.source_file}，发票号码 {invoice_no}，"
                            f"{duplicate_position}本次未写入；请查看 Excel 的“识别提示”页。"
                        )
                    for weak_row, existing_row in weak_duplicates:
                        notice = duplicate_notice_from_ledger_row(weak_row, existing_row).model_copy(
                            update={
                                "severity": "已写入",
                                "issue_type": "疑似重复（弱身份票）",
                                "action": "已写入采集表；请人工确认是否真重复，重复则删除该行。",
                            }
                        )
                        if notice.notice_id in existing_notice_ids:
                            continue
                        recognition_notices.append(notice)
                        existing_notice_ids.add(notice.notice_id)
                        duplicate_position = (
                            f"重复位置：采集表第 {existing_row['excel_row']} 行；"
                            if existing_row and existing_row.get("excel_row")
                            else ""
                        )
                        invoice_no = notice.invoice_no or "发票号码未识别"
                        result.messages.append(
                            f"疑似重复（弱身份票）：文件 {notice.source_file}，发票号码 {invoice_no}，"
                            f"{duplicate_position}本次已写入但需人工确认；请查看 Excel 的“识别提示”页。"
                        )
                result.skipped_duplicate_rows = skipped
                result.added_rows = written
            elif mode == "invoice_summary":
                written, skipped, _, _ = _append_rows(ws, fields, _summary_rows(ledger_rows))
            elif mode == "review_issues":
                if recognition_notices is None:
                    issues = _issue_rows(ledger_rows)
                    written, skipped, _, _ = _append_rows(ws, fields, issues)
                    result.review_required_rows = len(issues)
                else:
                    written, skipped = _append_notice_rows(ws, fields, recognition_notices)
                    result.review_required_rows = sum(
                        notice.issue_type == "需复核" for notice in recognition_notices
                    )
            result.actions.append(
                {
                    "action": WriteAction.ADDED.value,
                    "sheet": sheet_spec["name"],
                    "mode": mode,
                    "rows": written,
                    "skipped_duplicate_rows": skipped,
                }
            )
            if mode == "ledger_rows":
                _autofit_collection_sheet(ws)
        descriptor, temporary_name = tempfile.mkstemp(
            dir=workbook_path.parent,
            prefix=f".{workbook_path.name}.",
            suffix=".tmp.xlsx",
        )
        os.close(descriptor)
        temporary_path = Path(temporary_name)
        workbook.save(temporary_path)
        workbook.close()
        with temporary_path.open("r+b") as saved_file:
            os.fsync(saved_file.fileno())
        saved_report = validate_template_workbook(temporary_path, profile)
        if saved_report.get("blocked_write") is True or saved_report["status"] != "passed":
            raise ValueError(
                {
                    "message": "Saved workbook does not match profile",
                    "template_drift_report": saved_report,
                }
            )
        os.replace(temporary_path, workbook_path)
        temporary_path = None
    finally:
        workbook.close()
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)
    return result


def write_with_template_profile(
    workbook_path: str | Path,
    template_profile_path: str | Path,
    ledger_rows: list[LedgerRow],
    run_id: str,
    recognition_notices: list[RecognitionNotice] | None = None,
) -> WriteResult:
    with _exclusive_workbook_lock(Path(workbook_path)):
        return _write_with_template_profile(
            workbook_path,
            template_profile_path,
            ledger_rows,
            run_id,
            recognition_notices,
        )
