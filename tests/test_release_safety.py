from pathlib import Path
from shutil import copy2
from tempfile import TemporaryDirectory
import errno
import os
import unittest
from unittest.mock import patch

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from invoice_ledger.cli import AUTO_OCR_CONFIG, _load_runtime_config
from invoice_ledger.contracts import (
    InvoiceUnit,
    LedgerRow,
    OcrResult,
    OcrStatus,
    RecognitionStatus,
)
from invoice_ledger.input_profile.ocr_adapter import run_ocr_batch
from invoice_ledger.output.template_writer import write_with_template_profile
from invoice_ledger.pipeline.unit_processor import process_invoice_input


PROJECT_ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = PROJECT_ROOT / "发票采集台账.xlsx"
PROFILE = PROJECT_ROOT / "config" / "template_profiles" / "current.yaml"


def ledger_row(**updates):
    values = {
        "draft_row_id": "row-1",
        "run_id": "test-run",
        "source_file": "invoice.pdf",
        "page_range": [1],
        "invoice_unit_id": "unit-1",
        "schema_id": "digital-vat-invoice",
        "variant_id": "standard",
        "invoice_key": "invoice-1",
        "invoice_line_key": "invoice-1-line-1",
        "line_no": 1,
        "processed_at": "2026-07-27T10:00:00",
        "invoice_no": "12345678901234567890",
        "seller_name": "正常销方",
        "item_name": "服务费",
        "line_amount": "100.00",
        "line_tax_amount": "6.00",
        "line_total_with_tax": "106.00",
        "invoice_amount_total": "100.00",
        "invoice_tax_total": "6.00",
        "invoice_total_with_tax": "106.00",
        "invoice_type": "电子发票",
        "recognition_status": RecognitionStatus.READY,
        "confidence": 1.0,
    }
    values.update(updates)
    return LedgerRow(**values)


class ReleaseSafetyTests(unittest.TestCase):
    def test_failed_save_keeps_original_ledger_unchanged(self):
        with TemporaryDirectory() as directory:
            ledger = Path(directory) / "ledger.xlsx"
            copy2(TEMPLATE, ledger)
            original = ledger.read_bytes()

            def fail_after_partial_write(_workbook, filename):
                Path(filename).write_bytes(b"partial workbook")
                raise OSError("simulated interrupted save")

            with patch.object(Workbook, "save", fail_after_partial_write):
                with self.assertRaisesRegex(OSError, "simulated interrupted save"):
                    write_with_template_profile(
                        workbook_path=ledger,
                        template_profile_path=PROFILE,
                        ledger_rows=[],
                        recognition_notices=[],
                        run_id="atomic-failure",
                    )

            self.assertEqual(original, ledger.read_bytes())

    def test_formula_like_invoice_text_is_stored_as_plain_text(self):
        with TemporaryDirectory() as directory:
            ledger = Path(directory) / "ledger.xlsx"
            copy2(TEMPLATE, ledger)

            write_with_template_profile(
                workbook_path=ledger,
                template_profile_path=PROFILE,
                ledger_rows=[ledger_row(seller_name="=1+1")],
                recognition_notices=[],
                run_id="formula-safety",
            )

            workbook = load_workbook(ledger, data_only=False)
            try:
                sheet = workbook["发票信息采集"]
                headers = {cell.value: cell.column for cell in sheet[1]}
                cell = sheet.cell(2, headers["销方名称"])
                self.assertEqual("=1+1", cell.value)
                self.assertEqual("s", cell.data_type)
            finally:
                workbook.close()

    def test_atomic_save_preserves_append_and_duplicate_skip_behavior(self):
        with TemporaryDirectory() as directory:
            ledger = Path(directory) / "ledger.xlsx"
            copy2(TEMPLATE, ledger)
            row = ledger_row()

            first = write_with_template_profile(
                ledger,
                PROFILE,
                [row],
                "first-write",
                recognition_notices=[],
            )
            second = write_with_template_profile(
                ledger,
                PROFILE,
                [row],
                "second-write",
                recognition_notices=[],
            )

            self.assertEqual(1, first.added_rows)
            self.assertEqual(0, second.added_rows)
            self.assertEqual(1, second.skipped_duplicate_rows)
            workbook = load_workbook(ledger, data_only=False)
            try:
                for sheet_name in ("发票信息采集", "发票基础信息"):
                    sheet = workbook[sheet_name]
                    headers = {cell.value: cell.column for cell in sheet[1]}
                    written_ids = [
                        sheet.cell(row, headers["票据ID"]).value
                        for row in range(2, sheet.max_row + 1)
                        if sheet.cell(row, headers["票据ID"]).value
                    ]
                    self.assertEqual(["invoice-1-line-1"], written_ids)
            finally:
                workbook.close()
            self.assertEqual([], list(ledger.parent.glob(f".{ledger.name}.*.tmp.xlsx")))

    def test_corrupt_pdf_becomes_failed_unit_instead_of_aborting_batch(self):
        with TemporaryDirectory() as directory:
            invalid_pdf = Path(directory) / "broken.pdf"
            invalid_pdf.write_bytes(b"not a pdf")

            result = process_invoice_input(
                invalid_pdf,
                {"ocr": {"enabled": False}},
                "bad-file-isolation",
                "2026-07-27T10:00:00",
            )

            self.assertEqual(1, len(result["unit_results"]))
            unit_result = result["unit_results"][0]
            self.assertEqual(RecognitionStatus.FAILED, unit_result["invoice_record"].quality.status)
            self.assertIn("broken.pdf", unit_result["invoice_record"].source.source_file)

    def test_programming_error_is_not_disguised_as_bad_input(self):
        with patch(
            "invoice_ledger.pipeline.unit_processor._process_invoice_input",
            side_effect=RuntimeError("simulated programming error"),
        ):
            with self.assertRaisesRegex(RuntimeError, "simulated programming error"):
                process_invoice_input(
                    Path("invoice.pdf"),
                    {},
                    "programming-error",
                    "2026-07-27T10:00:00",
                )

    def test_workbook_lock_serializes_writers(self):
        from threading import Barrier, Lock, Thread
        from time import sleep
        from invoice_ledger.output import template_writer

        self.assertTrue(hasattr(template_writer, "_exclusive_workbook_lock"))
        barrier = Barrier(2)
        state_lock = Lock()
        active = 0
        maximum_active = 0
        errors = []

        def hold_lock():
            nonlocal active, maximum_active
            barrier.wait()
            try:
                with template_writer._exclusive_workbook_lock(TEMPLATE):
                    with state_lock:
                        active += 1
                        maximum_active = max(maximum_active, active)
                    sleep(0.05)
                    with state_lock:
                        active -= 1
            except Exception as exc:
                errors.append(exc)

        threads = [Thread(target=hold_lock) for _ in range(2)]
        for thread in threads:
            thread.start()
        for thread in threads:
            thread.join()

        self.assertEqual([], errors)
        self.assertEqual(1, maximum_active)

    @unittest.skipUnless(os.name == "nt", "Windows file-lock retry behavior")
    def test_workbook_lock_keeps_waiting_after_windows_retry_window(self):
        import msvcrt
        from invoice_ledger.output import template_writer

        lock_busy = OSError(errno.EDEADLK, "simulated lock retry timeout")
        with patch.object(msvcrt, "locking", side_effect=[lock_busy, None, None]) as locking:
            with template_writer._exclusive_workbook_lock(TEMPLATE):
                pass

        self.assertEqual(3, locking.call_count)

    def test_auto_gpu_ocr_failure_retries_on_cpu(self):
        with patch("invoice_ledger.cli._has_nvidia_gpu", return_value=True):
            runtime_config = _load_runtime_config(AUTO_OCR_CONFIG)
        self.assertEqual("cpu", runtime_config["ocr"].get("fallback_device"))

        unit = InvoiceUnit(
            invoice_unit_id="ocr-unit",
            source_file="invoice.png",
            page_range=[1],
            unit_type="image",
            status=RecognitionStatus.READY,
        )
        devices = []

        def fake_batch(units, ocr_config, pdf_context=None):
            devices.append(ocr_config["device"])
            status = OcrStatus.FAILED if ocr_config["device"].startswith("gpu") else OcrStatus.READY
            message = ["PaddleOCR batch failed on gpu: simulated"] if status == OcrStatus.FAILED else []
            return {
                unit.invoice_unit_id: OcrResult(
                    invoice_unit_id=unit.invoice_unit_id,
                    status=status,
                    provider="paddle",
                    source_file=unit.source_file,
                    page_range=unit.page_range,
                    messages=message,
                )
                for unit in units
            }

        with patch("invoice_ledger.input_profile.ocr_adapter._run_paddle_batch", fake_batch):
            result = run_ocr_batch([unit], runtime_config)

        self.assertEqual(["gpu:0", "cpu"], devices)
        self.assertEqual(OcrStatus.READY, result["ocr-unit"].status)


if __name__ == "__main__":
    unittest.main()
