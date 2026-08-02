from pathlib import Path
from shutil import copy2
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from invoice_ledger.contracts import (
    InvoiceFields,
    InvoiceQuality,
    InvoiceRecord,
    InvoiceSource,
    InvoiceUnit,
    LedgerRow,
    RecognitionStatus,
)
from invoice_ledger.output.recognition_notices import build_recognition_notices
from invoice_ledger.output.template_writer import write_with_template_profile
from invoice_ledger.validation.review_notes import user_review_remark


PROJECT_ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = PROJECT_ROOT / "发票采集台账.xlsx"
PROFILE = PROJECT_ROOT / "config" / "template_profiles" / "current.yaml"


def review_row(review_remark: str) -> LedgerRow:
    return LedgerRow(
        draft_row_id="row-1",
        run_id="test-run",
        source_file=r"D:\发票\测试发票.pdf",
        page_range=[1],
        invoice_unit_id="unit-1",
        invoice_key="invoice-1",
        invoice_line_key="invoice-1-line-1",
        line_no=1,
        processed_at="2026-08-03T10:00:00",
        invoice_no=None,
        invoice_total_with_tax="106.00",
        recognition_status=RecognitionStatus.REVIEW_REQUIRED,
        review_remark=review_remark,
    )


class RecognitionNoticeTests(unittest.TestCase):
    def test_review_notice_names_the_problem_and_keeps_only_clear_actions(self):
        notices = build_recognition_notices(
            [],
            [
                review_row(
                    "missing invoice_no; low confidence total_with_tax; "
                    "amount_total + tax_total != total_with_tax; conflict seller_name; "
                    "internal_unknown_flag"
                )
            ],
        )

        self.assertEqual(1, len(notices))
        self.assertEqual("缺少发票号码", notices[0].issue_type)
        self.assertEqual(
            "请对照原票补充发票号码或数电发票号码；"
            "请核对票面金额、税额和价税合计，三者对不上；"
            "请对照原票确认销售方名称，当前识别结果有冲突。",
            notices[0].action,
        )
        for machine_text in ("invoice_no", "total_with_tax", "confidence", "OCR", "internal"):
            self.assertNotIn(machine_text, notices[0].action)

    def test_one_invoice_keeps_one_review_notice_instead_of_one_per_line(self):
        first = review_row("missing invoice_no")
        second = first.model_copy(
            update={
                "draft_row_id": "row-2",
                "invoice_line_key": "invoice-1-line-2",
                "line_no": 2,
            }
        )

        notices = build_recognition_notices([], [first, second])

        self.assertEqual(1, len(notices))
        self.assertEqual("缺少发票号码", notices[0].issue_type)

    def test_failed_notice_hides_internal_error_and_tells_user_what_to_do(self):
        unit = InvoiceUnit(
            invoice_unit_id="failed-unit",
            source_file=r"D:\发票\模糊图片.pdf",
            page_range=[2],
            unit_type="pdf_page",
            status=RecognitionStatus.READY,
        )
        record = InvoiceRecord(
            invoice_unit_id=unit.invoice_unit_id,
            schema_id=None,
            variant_id=None,
            source=InvoiceSource(source_file=unit.source_file, page_range=[2]),
            invoice=InvoiceFields(),
            quality=InvoiceQuality(
                status=RecognitionStatus.FAILED,
                remark="parser_error: missing text_units at stage extract_content",
            ),
        )

        notices = build_recognition_notices(
            [{"invoice_unit": unit, "invoice_record": record}],
            [],
        )

        self.assertEqual("识别失败", notices[0].issue_type)
        self.assertEqual(
            "请确认原文件内容完整、可以正常打开；仍无法识别时，请手工补录。",
            notices[0].action,
        )
        self.assertNotIn("parser_error", notices[0].action)
        self.assertNotIn("text_units", notices[0].action)

    def test_already_human_review_text_stays_clear_in_notice(self):
        human_text = user_review_remark(
            "missing invoice_no; amount_total + tax_total != total_with_tax"
        )

        notice = build_recognition_notices([], [review_row(human_text)])[0]

        self.assertEqual("缺少发票号码", notice.issue_type)
        self.assertEqual(human_text, notice.action)

    def test_excel_keeps_specific_problem_and_counts_it_as_review_required(self):
        row = review_row("missing invoice_no")
        notices = build_recognition_notices([], [row])

        with TemporaryDirectory() as directory:
            ledger = Path(directory) / "ledger.xlsx"
            copy2(TEMPLATE, ledger)
            result = write_with_template_profile(
                workbook_path=ledger,
                template_profile_path=PROFILE,
                ledger_rows=[row],
                recognition_notices=notices,
                run_id="human-notice",
            )

            self.assertEqual(1, result.review_required_rows)
            workbook = load_workbook(ledger, data_only=True)
            try:
                sheet = workbook["识别提示"]
                headers = {cell.value: cell.column for cell in sheet[1]}
                self.assertEqual("缺少发票号码", sheet.cell(2, headers["问题"]).value)
                self.assertEqual(
                    "请对照原票补充发票号码或数电发票号码。",
                    sheet.cell(2, headers["处理方式"]).value,
                )
            finally:
                workbook.close()

    def test_amount_and_negative_value_prompts_keep_specific_problem_names(self):
        cases = [
            (
                "line 2: quantity * unit_price != line_amount",
                "金额不一致",
                "请核对第 2 行的数量、单价和金额，三者对不上。",
            ),
            (
                "待复核：可抵扣税额规则遇到负数项目 退票费",
                "负数金额需确认",
                "请确认“退票费”的负数金额是否与原票一致。",
            ),
        ]

        for raw, expected_issue, expected_action in cases:
            with self.subTest(raw=raw):
                action = user_review_remark(raw)
                self.assertEqual(expected_issue, build_recognition_notices([], [review_row(action)])[0].issue_type)
                self.assertEqual(expected_action, action)

    def test_reconciliation_prompt_states_the_amount_difference_plainly(self):
        raw = (
            "待复核：汇总金额与明细金额合计不一致，"
            "汇总金额 100.00，明细合计 90.00，差额 -10.00"
        )

        action = user_review_remark(raw)
        notice = build_recognition_notices([], [review_row(action)])[0]

        self.assertEqual("金额不一致", notice.issue_type)
        self.assertEqual("请核对票面金额和明细金额合计，二者相差 10.00 元。", action)

    def test_unclear_and_conflicting_fields_are_explained_without_machine_terms(self):
        cases = [
            (
                "low confidence total_with_tax",
                "识别不清",
                "请对照原票核对价税合计，图片可能识别不清。",
            ),
            (
                "conflict seller_name",
                "信息有冲突",
                "请对照原票确认销售方名称，当前识别结果有冲突。",
            ),
        ]

        for raw, expected_issue, expected_action in cases:
            with self.subTest(raw=raw):
                action = user_review_remark(raw)
                notice = build_recognition_notices([], [review_row(action)])[0]
                self.assertEqual(expected_issue, notice.issue_type)
                self.assertEqual(expected_action, action)


if __name__ == "__main__":
    unittest.main()
