from __future__ import annotations

import os
import re
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

from invoice_ledger.output.workbook_repair import (
    make_loadable_copy,
    repair_styles_xml,
    safe_load_workbook,
)


def _write_workbook() -> Path:
    descriptor, name = tempfile.mkstemp(suffix=".xlsx")
    os.close(descriptor)
    path = Path(name)
    wb = Workbook()
    wb.active["A1"] = "测试数据"
    wb.save(path)
    return path


def _corrupt_with_empty_fill(source: Path) -> Path:
    """复制并把第一个 <fill> 改成空 <fill/>,模拟 WPS 等工具的坏写法。"""
    with zipfile.ZipFile(source) as archive:
        items = {name: archive.read(name) for name in archive.namelist()}
    styles = items["xl/styles.xml"].decode("utf-8")
    styles = re.sub(r"<fill>.*?</fill>", "<fill/>", styles, count=1, flags=re.S)
    items["xl/styles.xml"] = styles.encode("utf-8")
    descriptor, name = tempfile.mkstemp(suffix=".xlsx")
    os.close(descriptor)
    corrupted = Path(name)
    with zipfile.ZipFile(corrupted, "w", zipfile.ZIP_DEFLATED) as archive:
        for entry, data in items.items():
            archive.writestr(entry, data)
    return corrupted


class WorkbookRepairTest(unittest.TestCase):
    def test_clean_styles_need_no_repair(self):
        path = _write_workbook()
        with zipfile.ZipFile(path) as archive:
            styles = archive.read("xl/styles.xml")
        self.assertIsNone(repair_styles_xml(styles))

    def test_empty_fill_breaks_plain_load(self):
        # 回归基准:空 fill 会让裸 load_workbook 崩(openpyxl issue #1805)
        corrupted = _corrupt_with_empty_fill(_write_workbook())
        with self.assertRaises(TypeError):
            load_workbook(corrupted)

    def test_safe_load_repairs_and_reads_data(self):
        corrupted = _corrupt_with_empty_fill(_write_workbook())
        wb = safe_load_workbook(corrupted)
        self.assertEqual(wb.active["A1"].value, "测试数据")

    def test_make_loadable_copy_preserves_clean_file(self):
        clean = _write_workbook()
        self.assertEqual(make_loadable_copy(clean), clean)


if __name__ == "__main__":
    unittest.main()
